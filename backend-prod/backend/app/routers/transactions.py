from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List, Literal, Tuple, Dict
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import uuid as _uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from pydantic import BaseModel, Field, field_validator
from enum import Enum

from app.routers.auth import get_current_user, UserDep
from app.database.postgresql import SessionLocal
from app.models.spendsense_models import (
    TxnFact, TxnEnriched, DimCategory, DimSubcategory
)
from sqlalchemy import func, text, and_, or_, select
from app.core.websocket_manager import websocket_manager

# =============================================================================
# Pydantic Models (with fixes from schema)
# =============================================================================

class TransactionType(str, Enum):
    DEBIT = "debit"
    CREDIT = "credit"
    CREDIT_CARD = "credit_card"  # treated as debit for direction

class TransactionStatus(str, Enum):
    PENDING = "pending"
    CLEARED = "cleared"
    FAILED = "failed"

class TransactionCreate(BaseModel):
    amount: Decimal = Field(..., ge=Decimal("0.01"), description="Transaction amount")
    currency: str = Field(default="INR", pattern="^[A-Z]{3}$", description="ISO 4217 currency code")
    transaction_date: datetime = Field(..., description="Transaction date and time (UTC preferred)")
    description: str = Field(..., min_length=1, max_length=1000)
    merchant: Optional[str] = Field(None, max_length=255)
    category: Optional[str] = Field(None, max_length=100)       # category_code
    subcategory: Optional[str] = Field(None, max_length=100)    # subcategory_code
    bank: Optional[str] = Field(None, max_length=100)
    account_type: Optional[str] = Field(None, max_length=50)
    transaction_type: TransactionType
    reference_id: Optional[str] = Field(None, max_length=255)
    tags: List[str] = Field(default_factory=list)  # NOTE: Pydantic doesn't enforce list max_length here
    status: TransactionStatus = Field(default=TransactionStatus.CLEARED)

    @field_validator('amount')
    @classmethod
    def v_amount(cls, v: Decimal) -> Decimal:
        # Force 2dp quantization to avoid floaty decimals
        q = v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if q <= 0:
            raise ValueError("Amount must be positive")
        if q > Decimal("999999999"):
            raise ValueError("Amount exceeds maximum allowed")
        return q

    @field_validator('currency')
    @classmethod
    def v_currency(cls, v: str) -> str:
        return v.upper()

    @field_validator('tags')
    @classmethod
    def v_tags(cls, v: List[str]) -> List[str]:
        # Enforce at-most 20 tags and trim/normalize
        if len(v) > 20:
            raise ValueError("At most 20 tags allowed")
        norm = []
        for t in v:
            t = (t or "").strip()
            if not t:
                continue
            if len(t) > 40:
                raise ValueError("Each tag must be <= 40 chars")
            norm.append(t)
        return norm

    class Config:
        json_schema_extra = {
            "example": {
                "amount": 2450.00,
                "currency": "INR",
                "transaction_date": "2024-01-15T10:30:00Z",
                "description": "Amazon purchase",
                "merchant": "AMAZON",
                "category": "shopping",
                "subcategory": "online_shopping",
                "bank": "HDFC",
                "transaction_type": "debit",
                "reference_id": "TXN123",
                "tags": ["online", "shopping"]
            }
        }

class TransactionResponse(BaseModel):
    id: str
    user_id: str
    amount: Decimal
    currency: str
    transaction_date: datetime
    description: str
    merchant: Optional[str]
    category: Optional[str]
    subcategory: Optional[str]
    bank: Optional[str]
    account_type: Optional[str]
    transaction_type: TransactionType
    reference_id: Optional[str]
    status: TransactionStatus
    tags: List[str]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

class TransactionListResponse(BaseModel):
    data: List[TransactionResponse]
    total: int
    skip: int
    limit: int

class TransactionStatsResponse(BaseModel):
    period: str
    total_debit: Decimal
    total_credit: Decimal
    net_amount: Decimal
    transaction_count: int
    average_transaction: Decimal
    top_category: Optional[str]
    top_merchant: Optional[str]

class TransactionUpdate(BaseModel):
    amount: Optional[Decimal] = None
    currency: Optional[str] = Field(None, pattern="^[A-Z]{3}$")
    transaction_date: Optional[datetime] = None
    description: Optional[str] = None
    merchant: Optional[str] = None
    category: Optional[str] = None
    subcategory: Optional[str] = None
    bank: Optional[str] = None
    transaction_type: Optional[TransactionType] = None
    reference_id: Optional[str] = None
    status: Optional[TransactionStatus] = None
    tags: Optional[List[str]] = None

    @field_validator('amount')
    @classmethod
    def u_amount(cls, v: Optional[Decimal]) -> Optional[Decimal]:
        if v is None:
            return v
        q = v.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if q <= 0:
            raise ValueError("Amount must be positive")
        return q

    @field_validator('currency')
    @classmethod
    def u_currency(cls, v: Optional[str]) -> Optional[str]:
        return v.upper() if v else v

class BulkImportError(BaseModel):
    row: int
    field: Optional[str] = None
    message: str

class BulkImportResponse(BaseModel):
    success: bool
    total_rows: int
    imported: int
    errors: List[BulkImportError]

# =============================================================================
# Router
# =============================================================================

router = APIRouter(tags=["Transactions"])  # Prefix handled in main.py


@router.get("/template")
async def download_transaction_template(user: UserDep = Depends(get_current_user)):
    session = SessionLocal()
    try:
        categories = session.query(DimCategory).filter(
            DimCategory.active.is_(True)
        ).order_by(
            DimCategory.display_order.asc(),
            DimCategory.category_name.asc()
        ).all()

        subcategories = session.query(DimSubcategory).filter(
            DimSubcategory.active.is_(True)
        ).order_by(
            DimSubcategory.display_order.asc(),
            DimSubcategory.subcategory_name.asc()
        ).all()

        categories_data = [
            {
                "category_code": cat.category_code,
                "category_name": cat.category_name,
                "txn_type": cat.txn_type
            }
            for cat in categories
        ]

        subcategories_data = [
            {
                "subcategory_code": sub.subcategory_code,
                "subcategory_name": sub.subcategory_name,
                "category_code": sub.category_code
            }
            for sub in subcategories
        ]
    finally:
        session.close()

    workbook = Workbook()
    transactions_sheet = workbook.active
    transactions_sheet.title = "Transactions"

    headers = [
        "merchant",
        "description",
        "amount",
        "transaction_date",
        "transaction_type",
        "category_code",
        "subcategory_code",
        "currency",
        "bank",
        "reference_id",
        "tags"
    ]
    transactions_sheet.append(headers)

    header_font = Font(bold=True)
    for cell in transactions_sheet[1]:
        cell.font = header_font

    sample_row = [
        "Example Merchant",
        "Sample description",
        2500.00,
        "2025-01-15T10:30:00Z",
        "debit",
        "groceries",
        "",
        "INR",
        "HDFC",
        "REF123",
        ""
    ]
    transactions_sheet.append(sample_row)
    transactions_sheet.freeze_panes = "A2"

    # Add data validation dropdowns
    # Transaction type dropdown (column E, index 5)
    transaction_type_validation = DataValidation(
        type="list",
        formula1="debit,credit",
        allow_blank=True
    )
    transaction_type_validation.add("E2:E10000")  # Apply to all rows starting from row 2
    transactions_sheet.add_data_validation(transaction_type_validation)

    # Category code dropdown (column F, index 6)
    category_codes = ",".join([cat["category_code"] for cat in categories_data])
    category_validation = DataValidation(
        type="list",
        formula1=category_codes,
        allow_blank=True
    )
    category_validation.add("F2:F10000")
    transactions_sheet.add_data_validation(category_validation)

    # Subcategory code dropdown (column G, index 7)
    subcategory_codes = ",".join([sub["subcategory_code"] for sub in subcategories_data if sub["subcategory_code"]])
    if subcategory_codes:  # Only add if there are subcategories
        subcategory_validation = DataValidation(
            type="list",
            formula1=subcategory_codes,
            allow_blank=True
        )
        subcategory_validation.add("G2:G10000")
        transactions_sheet.add_data_validation(subcategory_validation)

    reference_sheet = workbook.create_sheet("Categories")
    reference_headers = [
        "category_code",
        "category_name",
        "subcategory_code",
        "subcategory_name"
    ]
    reference_sheet.append(reference_headers)
    for cell in reference_sheet[1]:
        cell.font = header_font

    subcategory_map: Dict[str, List[Dict[str, str]]] = {}
    for sub in subcategories_data:
        subcategory_map.setdefault(sub["category_code"], []).append(sub)

    for cat in categories_data:
        entries = subcategory_map.get(cat["category_code"]) or [None]
        for sub in entries:
            reference_sheet.append([
                cat["category_code"],
                cat["category_name"],
                sub["subcategory_code"] if sub else "",
                sub["subcategory_name"] if sub else ""
            ])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"transaction_template_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.post("/bulk-import", response_model=BulkImportResponse)
async def bulk_import_transactions(
    file: UploadFile = File(...),
    user: UserDep = Depends(get_current_user)
):
    """
    Bulk import transactions from Excel (.xlsx) or CSV file.
    Validates all rows before importing. Fails entire upload if any row has errors.
    """
    # Validate file type
    file_ext = file.filename.lower() if file.filename else ""
    if not (file_ext.endswith('.xlsx') or file_ext.endswith('.csv')):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx or .csv files are allowed"
        )
    
    # Read file content
    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File is empty")
    
    # Parse file based on extension
    try:
        if file_ext.endswith('.xlsx'):
            df = pd.read_excel(BytesIO(content), sheet_name=0, engine='openpyxl')
        else:  # CSV
            df = pd.read_csv(BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    if df.empty:
        raise HTTPException(status_code=400, detail="File contains no data rows")
    
    # Expected columns (case-insensitive matching)
    expected_cols = [
        'merchant', 'description', 'amount', 'transaction_date',
        'transaction_type', 'category_code', 'subcategory_code',
        'currency', 'bank', 'reference_id', 'tags'
    ]
    
    # Normalize column names (lowercase, strip whitespace)
    df.columns = df.columns.str.lower().str.strip()
    
    # Check required columns
    required_cols = ['description', 'amount', 'transaction_date', 'transaction_type']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing_cols)}"
        )
    
    session = SessionLocal()
    errors: List[BulkImportError] = []
    validated_rows: List[Dict] = []
    
    try:
        # Validate all rows first
        for idx, row in df.iterrows():
            row_num = idx + 2  # +2 because Excel rows start at 1 and we skip header
            
            try:
                # Extract and validate fields
                merchant = str(row.get('merchant', '')).strip() if pd.notna(row.get('merchant')) else ''
                description = str(row.get('description', '')).strip() if pd.notna(row.get('description')) else ''
                amount_str = row.get('amount', '')
                date_str = row.get('transaction_date', '')
                txn_type_str = str(row.get('transaction_type', '')).strip().lower() if pd.notna(row.get('transaction_type')) else ''
                category_code = str(row.get('category_code', '')).strip() if pd.notna(row.get('category_code')) else None
                subcategory_code = str(row.get('subcategory_code', '')).strip() if pd.notna(row.get('subcategory_code')) else None
                currency = str(row.get('currency', 'INR')).strip().upper() if pd.notna(row.get('currency')) else 'INR'
                bank = str(row.get('bank', '')).strip() if pd.notna(row.get('bank')) else None
                reference_id = str(row.get('reference_id', '')).strip() if pd.notna(row.get('reference_id')) else None
                tags_str = str(row.get('tags', '')).strip() if pd.notna(row.get('tags')) else ''
                
                # Validate required fields
                if not description:
                    errors.append(BulkImportError(row=row_num, field='description', message='Description is required'))
                    continue
                
                if pd.isna(amount_str) or amount_str == '':
                    errors.append(BulkImportError(row=row_num, field='amount', message='Amount is required'))
                    continue
                
                try:
                    amount = Decimal(str(amount_str))
                    if amount <= 0:
                        errors.append(BulkImportError(row=row_num, field='amount', message='Amount must be positive'))
                        continue
                    amount = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                except (ValueError, TypeError):
                    errors.append(BulkImportError(row=row_num, field='amount', message=f'Invalid amount: {amount_str}'))
                    continue
                
                if pd.isna(date_str) or date_str == '':
                    errors.append(BulkImportError(row=row_num, field='transaction_date', message='Transaction date is required'))
                    continue
                
                try:
                    # Try parsing as datetime string (ISO format or common formats)
                    if isinstance(date_str, str):
                        # Try ISO format first
                        try:
                            txn_date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                        except:
                            # Try pandas parsing
                            txn_date = pd.to_datetime(date_str)
                            if isinstance(txn_date, pd.Timestamp):
                                txn_date = txn_date.to_pydatetime()
                    else:
                        txn_date = pd.to_datetime(date_str).to_pydatetime()
                except Exception as e:
                    errors.append(BulkImportError(row=row_num, field='transaction_date', message=f'Invalid date format: {date_str}'))
                    continue
                
                if not txn_type_str or txn_type_str not in ['debit', 'credit']:
                    errors.append(BulkImportError(row=row_num, field='transaction_type', message='Transaction type must be "debit" or "credit"'))
                    continue
                
                transaction_type = TransactionType.DEBIT if txn_type_str == 'debit' else TransactionType.CREDIT
                
                # Validate currency
                if len(currency) != 3 or not currency.isalpha():
                    errors.append(BulkImportError(row=row_num, field='currency', message='Currency must be a 3-letter ISO code'))
                    continue
                
                # Validate category/subcategory if provided
                if category_code or subcategory_code:
                    try:
                        cat_code, subcat_code = _ensure_category_pair(session, category_code, subcategory_code)
                    except HTTPException as e:
                        errors.append(BulkImportError(row=row_num, field='category_code' if category_code else 'subcategory_code', message=e.detail))
                        continue
                else:
                    cat_code, subcat_code = None, None
                
                # Parse tags if provided
                tags = []
                if tags_str:
                    tags = [t.strip() for t in tags_str.split(',') if t.strip()]
                    if len(tags) > 20:
                        errors.append(BulkImportError(row=row_num, field='tags', message='Maximum 20 tags allowed'))
                        continue
                
                # Create TransactionCreate object for validation
                try:
                    txn_create = TransactionCreate(
                        amount=amount,
                        currency=currency,
                        transaction_date=txn_date,
                        description=description,
                        merchant=merchant if merchant else None,
                        category=cat_code,
                        subcategory=subcat_code,
                        bank=bank,
                        transaction_type=transaction_type,
                        reference_id=reference_id,
                        tags=tags
                    )
                    validated_rows.append({
                        'txn_create': txn_create,
                        'row_num': row_num
                    })
                except Exception as e:
                    errors.append(BulkImportError(row=row_num, field=None, message=f'Validation error: {str(e)}'))
                    continue
                    
            except Exception as e:
                errors.append(BulkImportError(row=row_num, field=None, message=f'Unexpected error: {str(e)}'))
                continue
        
        # If any errors, return them all without importing
        if errors:
            return BulkImportResponse(
                success=False,
                total_rows=len(df),
                imported=0,
                errors=errors
            )
        
        # All rows validated - now import them
        from app.models.spendsense_models import UploadBatch
        user_id_uuid = _uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id
        
        # Get or create manual upload batch
        manual_upload = session.query(UploadBatch).filter(
            UploadBatch.source_type == 'manual',
            UploadBatch.user_id == user_id_uuid
        ).first()
        
        if not manual_upload:
            manual_upload = UploadBatch(
                user_id=user_id_uuid,
                source_type='manual',
                file_name=file.filename or 'bulk_import',
                status='loaded',
                total_records=0,
                parsed_records=0
            )
            session.add(manual_upload)
            session.flush()
        
        imported_count = 0
        for row_data in validated_rows:
            txn_create = row_data['txn_create']
            
            # Validate category/subcategory again (in case DB changed)
            cat_code, subcat_code = _ensure_category_pair(session, txn_create.category, txn_create.subcategory)
            direction = _direction_from_type(txn_create.transaction_type)
            
            txn = TxnFact(
                user_id=user_id_uuid,
                upload_id=manual_upload.upload_id,
                source_type='manual',
                account_ref=txn_create.bank,
                txn_external_id=txn_create.reference_id,
                txn_date=txn_create.transaction_date.date(),
                description=txn_create.description,
                amount=txn_create.amount,
                direction=direction,
                currency=txn_create.currency,
                merchant_name_norm=txn_create.merchant,
            )
            
            session.add(txn)
            session.flush()
            
            if cat_code or subcat_code:
                session.add(TxnEnriched(
                    txn_id=txn.txn_id,
                    category_code=cat_code,
                    subcategory_code=subcat_code,
                    rule_confidence=Decimal("0.90")
                ))
            
            # Try to compute dedupe fingerprint
            try:
                session.execute(text("""
                    UPDATE spendsense.txn_fact 
                    SET dedupe_fp = spendsense.fn_txn_fact_fp(
                        :user_id, :txn_date, :amount, :direction, 
                        :description, :merchant, :account_ref
                    ) 
                    WHERE txn_id = :txn_id
                """), {
                    "txn_id": str(txn.txn_id),
                    "user_id": str(user_id_uuid),
                    "txn_date": txn.txn_date,
                    "amount": float(txn.amount),
                    "direction": txn.direction,
                    "description": txn.description or "",
                    "merchant": txn.merchant_name_norm or "",
                    "account_ref": txn.account_ref or ""
                })
            except Exception:
                pass  # Non-fatal
            
            imported_count += 1
        
        session.commit()
        
        # Broadcast to websocket clients
        try:
            await websocket_manager.broadcast_to_user(user.user_id, {
                "type": "transactions_bulk_imported",
                "data": {"count": imported_count}
            })
        except Exception:
            pass  # Non-fatal
        
        return BulkImportResponse(
            success=True,
            total_rows=len(df),
            imported=imported_count,
            errors=[]
        )
        
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to import transactions: {str(e)}")
    finally:
        session.close()


# -----------------------
# Helpers
# -----------------------

def _direction_from_type(t: TransactionType) -> str:
    # CREDIT_CARD spends → treat as debit
    if t == TransactionType.CREDIT:
        return "credit"
    return "debit"

def _ensure_category_pair(session, category_code: Optional[str], subcategory_code: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """
    Validates that category/subcategory exist and align.
    Returns normalized (category_code, subcategory_code).
    """
    if not category_code and not subcategory_code:
        return None, None

    cat = None
    sub = None

    if category_code:
        cat = session.query(DimCategory).filter(
            DimCategory.category_code == category_code,
            DimCategory.active.is_(True)
        ).first()
        if not cat:
            raise HTTPException(status_code=400, detail=f"Unknown category_code '{category_code}'")

    if subcategory_code:
        sub = session.query(DimSubcategory).filter(
            DimSubcategory.subcategory_code == subcategory_code,
            DimSubcategory.active.is_(True)
        ).first()
        if not sub:
            raise HTTPException(status_code=400, detail=f"Unknown subcategory_code '{subcategory_code}'")

    if sub and cat and sub.category_code != cat.category_code:
        raise HTTPException(status_code=400, detail=f"subcategory_code '{subcategory_code}' does not belong to category_code '{category_code}'")

    # If only subcategory given, backfill its category
    if sub and not cat:
        cat = session.query(DimCategory).filter(
            DimCategory.category_code == sub.category_code
        ).first()

    return (cat.category_code if cat else None, sub.subcategory_code if sub else None)

def _to_response(txn: TxnFact, enriched: Optional[TxnEnriched]) -> TransactionResponse:
    # Map DB → API
    # Determine transaction_type from direction
    if txn.direction == "credit":
        txn_type = TransactionType.CREDIT
    else:
        txn_type = TransactionType.DEBIT
    
    return TransactionResponse(
        id=str(txn.txn_id),
        user_id=str(txn.user_id),
        amount=txn.amount,
        currency=txn.currency or "INR",
        transaction_date=datetime.combine(txn.txn_date, datetime.min.time()),
        description=txn.description or "",
        merchant=txn.merchant_name_norm,
        category=(enriched.category_code if enriched else None),
        subcategory=(enriched.subcategory_code if enriched else None),
        bank=txn.account_ref,                 # using account_ref as a "bank" hint
        account_type=None,                    # not modeled in fact; keep None unless you add it
        transaction_type=txn_type,
        reference_id=txn.txn_external_id,
        status=TransactionStatus.CLEARED,     # fact doesn't have explicit status; default cleared
        tags=[],                               # tags not modeled; extend with a join to a tags table if you add one
        created_at=txn.created_at,
        updated_at=txn.created_at,            # no updated_at in fact; mirror created_at
    )

# -----------------------
# Create
# -----------------------

@router.post("/", response_model=TransactionResponse, status_code=201)
async def create_transaction(payload: TransactionCreate, user: UserDep = Depends(get_current_user)):
    session = SessionLocal()
    try:
        # Validate & align category/subcategory
        cat_code, subcat_code = _ensure_category_pair(session, payload.category, payload.subcategory)

        # Direction from enum
        direction = _direction_from_type(payload.transaction_type)

        # Build TxnFact
        # Create a minimal upload_batch for manual transactions if needed
        # For now, we'll use a placeholder UUID that represents "manual" transactions
        # In production, you might want to create a system upload_batch for manual entries
        from app.models.spendsense_models import UploadBatch
        manual_upload = session.query(UploadBatch).filter(
            UploadBatch.source_type == 'manual',
            UploadBatch.user_id == _uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id
        ).first()
        
        if not manual_upload:
            # Create a system upload batch for manual transactions
            manual_upload = UploadBatch(
                user_id=_uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id,
                source_type='manual',
                file_name='manual_entry',
                status='loaded',
                total_records=0,
                parsed_records=0
            )
            session.add(manual_upload)
            session.flush()
        
        txn = TxnFact(
            user_id=_uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id,
            upload_id=manual_upload.upload_id,
            source_type='manual',
            account_ref=payload.bank,            # store bank hint here
            txn_external_id=payload.reference_id,
            txn_date=payload.transaction_date.date(),
            description=payload.description,
            amount=payload.amount,
            direction=direction,
            currency=payload.currency,
            merchant_name_norm=(payload.merchant or None),
        )

        # Persist
        session.add(txn)
        session.flush()  # acquire txn_id

        # Optional enrichment row
        if cat_code or subcat_code:
            session.add(TxnEnriched(
                txn_id=txn.txn_id,
                category_code=cat_code,
                subcategory_code=subcat_code,
                rule_confidence=Decimal("0.90")
            ))

        # Try to compute and store dedupe fingerprint if your schema has a column & function
        # (Assumes you added dedupe_fp and fn_txn_fact_fp() that takes user_id, date, amount, direction, description, merchant, account_ref)
        try:
            session.execute(text("""
                UPDATE spendsense.txn_fact 
                SET dedupe_fp = spendsense.fn_txn_fact_fp(
                    :user_id, :txn_date, :amount, :direction, 
                    :description, :merchant, :account_ref
                ) 
                WHERE txn_id = :txn_id
            """), {
                "txn_id": str(txn.txn_id),
                "user_id": str(txn.user_id),
                "txn_date": txn.txn_date,
                "amount": float(txn.amount),
                "direction": txn.direction,
                "description": txn.description or "",
                "merchant": txn.merchant_name_norm or "",
                "account_ref": txn.account_ref or ""
            })
        except Exception as fp_err:
            # Non-fatal if function/column not present
            print(f"⚠️  Dedupe fingerprint update skipped: {fp_err}")
            pass

        session.commit()

        # Map response
        enriched = session.query(TxnEnriched).filter(TxnEnriched.txn_id == txn.txn_id).first()
        response = _to_response(txn, enriched)

        # Broadcast to websocket clients
        try:
            await websocket_manager.broadcast_to_user(user.user_id, {
                "type": "transaction_created",
                "data": response.dict()
            })
        except Exception as ws_err:
            # Don't fail the request if websocket fails
            print(f"⚠️  WebSocket broadcast failed: {ws_err}")

        return response

    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        # Handle unique dedupe violations gracefully
        err = str(e).lower()
        if "ux_txn_fact_dedupe_fp" in err or "duplicate" in err or "unique constraint" in err:
            raise HTTPException(status_code=409, detail="Duplicate transaction")
        raise HTTPException(status_code=400, detail=f"Failed to create: {e}")
    finally:
        session.close()

# -----------------------
# Read (list)
# -----------------------

@router.get("/", response_model=TransactionListResponse)
async def list_transactions(
    user: UserDep = Depends(get_current_user),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    category: Optional[str] = None,
    subcategory: Optional[str] = None,
    direction: Optional[Literal["debit", "credit"]] = None,
    sort: Optional[Literal["date_desc","date_asc","amt_desc","amt_asc"]] = "date_desc",
    search: Optional[str] = None
):
    session = SessionLocal()
    try:
        user_id = _uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id

        q = session.query(TxnFact, TxnEnriched).outerjoin(TxnEnriched, TxnFact.txn_id == TxnEnriched.txn_id)\
             .filter(TxnFact.user_id == user_id)

        if start_date:
            q = q.filter(TxnFact.txn_date >= start_date)
        if end_date:
            q = q.filter(TxnFact.txn_date <= end_date)
        if category:
            q = q.filter(TxnEnriched.category_code == category)
        if subcategory:
            q = q.filter(TxnEnriched.subcategory_code == subcategory)
        if direction:
            q = q.filter(TxnFact.direction == direction)
        if search:
            s = f"%{search.strip().lower()}%"
            q = q.filter(or_(func.lower(TxnFact.description).like(s),
                             func.lower(TxnFact.merchant_name_norm).like(s)))

        total = q.count()

        if sort == "date_desc":
            q = q.order_by(TxnFact.txn_date.desc(), TxnFact.created_at.desc())
        elif sort == "date_asc":
            q = q.order_by(TxnFact.txn_date.asc(), TxnFact.created_at.asc())
        elif sort == "amt_desc":
            q = q.order_by(TxnFact.amount.desc())
        elif sort == "amt_asc":
            q = q.order_by(TxnFact.amount.asc())

        rows = q.offset(skip).limit(limit).all()

        data: List[TransactionResponse] = []
        for f, e in rows:
            data.append(_to_response(f, e))

        return TransactionListResponse(
            data=data,
            total=total,
            skip=skip,
            limit=limit
        )
    finally:
        session.close()

# -----------------------
# Read (single)
# -----------------------

@router.get("/{txn_id}", response_model=TransactionResponse)
async def get_transaction(txn_id: str, user: UserDep = Depends(get_current_user)):
    session = SessionLocal()
    try:
        uid = _uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id
        f = session.query(TxnFact).filter(TxnFact.txn_id == _uuid.UUID(txn_id), TxnFact.user_id == uid).first()
        if not f:
            raise HTTPException(status_code=404, detail="Transaction not found")
        e = session.query(TxnEnriched).filter(TxnEnriched.txn_id == f.txn_id).first()
        return _to_response(f, e)
    finally:
        session.close()

# -----------------------
# Update (partial)
# -----------------------

@router.patch("/{txn_id}", response_model=TransactionResponse)
async def update_transaction(txn_id: str, payload: TransactionUpdate, user: UserDep = Depends(get_current_user)):
    session = SessionLocal()
    try:
        uid = _uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id
        f = session.query(TxnFact).filter(TxnFact.txn_id == _uuid.UUID(txn_id), TxnFact.user_id == uid).first()
        if not f:
            raise HTTPException(status_code=404, detail="Transaction not found")

        # Apply scalar updates
        if payload.amount is not None: f.amount = payload.amount
        if payload.currency is not None: f.currency = payload.currency
        if payload.transaction_date is not None: f.txn_date = payload.transaction_date.date()
        if payload.description is not None: f.description = payload.description
        if payload.merchant is not None: f.merchant_name_norm = payload.merchant or None
        if payload.bank is not None: f.account_ref = payload.bank or None
        if payload.reference_id is not None: f.txn_external_id = payload.reference_id or None
        if payload.transaction_type is not None: f.direction = _direction_from_type(payload.transaction_type)

        # Upsert enrichment if category/subcategory provided
        # Use high confidence (0.99) for manual edits to prevent re-enrichment from overriding
        if payload.category is not None or payload.subcategory is not None:
            cat_code, subcat_code = _ensure_category_pair(session, payload.category, payload.subcategory)
            e = session.query(TxnEnriched).filter(TxnEnriched.txn_id == f.txn_id).first()
            if e:
                e.category_code = cat_code
                e.subcategory_code = subcat_code
                e.rule_confidence = Decimal("0.99")  # High confidence prevents auto re-enrichment
                # Clear matched_rule_id to indicate this is a manual override
                e.matched_rule_id = None
            else:
                session.add(TxnEnriched(
                    txn_id=f.txn_id,
                    category_code=cat_code,
                    subcategory_code=subcat_code,
                    rule_confidence=Decimal("0.99"),  # High confidence for manual edits
                    matched_rule_id=None  # Manual override, not from a rule
                ))

        session.commit()
        
        # Learn from this edit: create/update merchant_rules automatically
        try:
            from app.services.learning_service import learn_from_edit
            
            merchant_name = payload.merchant if payload.merchant is not None else f.merchant_name_norm
            description = payload.description if payload.description is not None else f.description
            cat_code, subcat_code = _ensure_category_pair(session, payload.category, payload.subcategory)
            
            if cat_code and (merchant_name or description):
                rule_id = learn_from_edit(
                    user_id=str(uid),
                    merchant_name=merchant_name,
                    description=description,
                    category_code=cat_code,
                    subcategory_code=subcat_code,
                    txn_id=str(f.txn_id)
                )
                if rule_id:
                    print(f"✅ Learned from edit: Created/updated rule {rule_id} for merchant='{merchant_name}', category='{cat_code}'")
        except Exception as learn_err:
            # Don't fail the update if learning fails
            print(f"⚠️  Warning: Failed to learn from edit: {learn_err}")

        e = session.query(TxnEnriched).filter(TxnEnriched.txn_id == f.txn_id).first()
        return _to_response(f, e)
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to update: {e}")
    finally:
        session.close()

# -----------------------
# Delete
# -----------------------

@router.delete("/{txn_id}", status_code=204)
async def delete_transaction(txn_id: str, user: UserDep = Depends(get_current_user)):
    session = SessionLocal()
    try:
        uid = _uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id
        f = session.query(TxnFact).filter(TxnFact.txn_id == _uuid.UUID(txn_id), TxnFact.user_id == uid).first()
        if not f:
            raise HTTPException(status_code=404, detail="Transaction not found")
        session.delete(f)
        session.commit()
        return
    finally:
        session.close()

# -----------------------
# Stats
# -----------------------

@router.get("/stats", response_model=TransactionStatsResponse)
async def transaction_stats(
    user: UserDep = Depends(get_current_user),
    period: Literal["day","week","month","year"] = Query("month")
):
    """
    Lightweight stats from fact + enriched.
    (For production dashboards, prefer your materialized views.)
    """
    session = SessionLocal()
    try:
        uid = _uuid.UUID(user.user_id) if isinstance(user.user_id, str) else user.user_id

        # Date window
        today = date.today()
        if period == "day":
            start = today
        elif period == "week":
            start = today.replace(day=max(1, today.day-6))
        elif period == "month":
            start = today.replace(day=1)
        else:  # year
            start = today.replace(month=1, day=1)

        q = session.query(TxnFact, TxnEnriched)\
            .outerjoin(TxnEnriched, TxnFact.txn_id == TxnEnriched.txn_id)\
            .filter(TxnFact.user_id == uid, TxnFact.txn_date >= start)

        rows = q.all()
        total_debit = Decimal("0")
        total_credit = Decimal("0")
        count = 0
        cat_counter = {}
        merch_counter = {}

        for f, e in rows:
            count += 1
            amt = f.amount or Decimal("0")
            if f.direction == "debit":
                total_debit += amt
            else:
                total_credit += amt

            # count categories (skip transfers to keep KPIs clean, optional)
            cat = e.category_code if e else None
            if cat and cat != "transfers":
                cat_counter[cat] = cat_counter.get(cat, 0) + float(amt)

            merch = (f.merchant_name_norm or "").strip().upper() if f.merchant_name_norm else None
            if merch:
                merch_counter[merch] = merch_counter.get(merch, 0) + float(amt)

        net = total_credit - total_debit
        avg = (total_debit + total_credit) / count if count else Decimal("0")

        top_cat = max(cat_counter, key=cat_counter.get) if cat_counter else None
        top_merch = max(merch_counter, key=merch_counter.get) if merch_counter else None

        return TransactionStatsResponse(
            period=period,
            total_debit=total_debit.quantize(Decimal("0.01")),
            total_credit=total_credit.quantize(Decimal("0.01")),
            net_amount=net.quantize(Decimal("0.01")),
            transaction_count=count,
            average_transaction=avg.quantize(Decimal("0.01")),
            top_category=top_cat,
            top_merchant=top_merch
        )
    finally:
        session.close()
